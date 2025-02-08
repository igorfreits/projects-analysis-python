import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook
import os

# Caminho dos arquivos
data_path = 'PROCESSADO ERRO/Analise de Dados/'
# ARQUIVOS
processado_erro = pd.read_excel(data_path + 'Processado Erro.xlsx')
realocacao = pd.read_excel(data_path + 'Realocacao.xlsx')

parametros, list_erros, info, clientes_fcm = [
    pd.read_excel(data_path + 'Parametros.xlsx', sheet_name=sheet)

    for sheet in ['Parametros', 'Lista de erros', 'Info', 'Clientes FCM']]

# DeclaraÃ§Ã£o de guias - RelatÃ³rio DASH
wb = load_workbook(data_path + 'Relatorio - Dash.xlsx')
relatorio_base = wb['Processado Erro - BASE']

# Limpeza de planilha
relatorio_base.delete_rows(2, relatorio_base.max_row)
relatorio_base.delete_cols(2, relatorio_base.max_column)

# CriaÃ§Ã£o de colunas
processado_erro['MÃªs AlteraÃ§Ã£o'] = ''
processado_erro['Dias Parados no Erro'] = ''
processado_erro['OBTS'] = ''
processado_erro['CAMPO'] = 'NÃ£o identificado'
processado_erro['ORIGEM DO ERRO'] = 'AnÃ¡lise Benner'
processado_erro['TIPO DE ERRO'] = 'Sistema'
processado_erro['CATEGORIA DE ERRO'] = 'Qualidade dos dados'
processado_erro['EMPRESA'] = ''
processado_erro['RESPONSÃVEL'] = 'OperaÃ§Ãµes CORP'

# Preenchimento de valores nulos
processado_erro['Mensagem Erro'].fillna(
    'Erro nÃ£o localizado', inplace=True)

processado_erro.fillna('-', inplace=True)

# InserÃ§Ã£o dos OBTS
processado_erro['OBTS'] = processado_erro['OBT']
processado_erro['OBTS'] = processado_erro['OBTS'].str.replace(
    'TMS', 'ARGO(TMS)')

mask_zupper = (processado_erro['Canal de Vendas'].str.contains('ZUPPER', case=False) |
               processado_erro['Grupo Empresarial'].str.contains('Zupper', case=False))

processado_erro['OBTS'] = processado_erro['OBTS'].mask(mask_zupper, 'ZUPPER')

mask_kontrip = processado_erro['Canal de Vendas'].str.contains(
    'KONTRIP', case=False)
processado_erro['OBTS'] = processado_erro['OBTS'].mask(mask_kontrip, 'KONTRIP')

processado_erro['Mensagem Erro'] = processado_erro['Mensagem Erro'].astype(str)
parametros['Mensagem'] = parametros['Mensagem'].astype(str)

# AtribuiÃ§Ã£o de campos, origem do erro e tipo de erro
for row in range(len(processado_erro)):
    for row2 in range(len(parametros)):
        if parametros['Mensagem'][row2] in processado_erro['Mensagem Erro'][row]:
            processado_erro.at[row, 'CAMPO'] = parametros.at[row2, 'Campo']
            processado_erro.at[row,
                               'ORIGEM DO ERRO'] = parametros.at[row2, 'Origem do Erro']
            processado_erro.at[row,
                               'TIPO DE ERRO'] = parametros.at[row2, 'Tipo de Erro']
            processado_erro.at[row,
                               'CATEGORIA DE ERRO'] = parametros.at[row2, 'Categoria de Erro']

# FormataÃ§Ã£o de Datas
processado_erro['Dias Parados no Erro'] = (
    datetime.now() - pd.to_datetime(processado_erro['Data InclusÃ£o'].str[:10], format='%d/%m/%Y')).dt.days

# Leitura de data de alteraÃ§Ã£o - EdiÃ§Ã£o nÃ£o permitida
processado_erro.loc[processado_erro['TIPO DE ERRO'].str.contains('EdiÃ§Ã£o nÃ£o Permitida'), 'Dias Parados no Erro'] = (
    datetime.now() -
    pd.to_datetime(
        processado_erro['Data AlteraÃ§Ã£o'].str[:10], format='%d/%m/%Y')
).dt.days

# Leitura de data de alteraÃ§Ã£o - Bilhete duplicado
processado_erro.loc[processado_erro['CAMPO'].str.contains('Bilhete duplicado'), 'Dias Parados no Erro'] = (
    datetime.now() -
    pd.to_datetime(
        processado_erro['Data AlteraÃ§Ã£o'].str[:10], format='%d/%m/%Y')
).dt.days

# MÃªs e ano de alteraÃ§Ã£o
processado_erro['MÃªs AlteraÃ§Ã£o'] = pd.to_datetime(
    processado_erro['Data InclusÃ£o'].str.split().str[0], format='%d/%m/%Y').dt.strftime('%B de %Y')

processado_erro.loc[processado_erro['Data EmissÃ£o'].str.contains(
    '-'), 'Data EmissÃ£o'] = processado_erro['Data InclusÃ£o']

# CategorizaÃ§Ã£o de dias parados
limites = [0, 3, 6, 9, 16, 24, 31, float('inf')]
rotulos = ['0 a 02 dias', '03 a 05 dias', '06 a 08 dias',
           '09 a 15 dias', '16 a 23 dias', '24 a 31 dias', '31 dias ou +']

processado_erro['Dias Parados no Erro'] = pd.cut(
    processado_erro['Dias Parados no Erro'], bins=limites, labels=rotulos, right=False, include_lowest=True)

# formataÃ§Ã£o de datas
# Aplicar formataÃ§Ã£o com dayfirst=True para ajustar o formato de data "DD/MM/AAAA HH:MM"
processado_erro['Data InclusÃ£o'] = pd.to_datetime(
    processado_erro['Data InclusÃ£o'], format='mixed', dayfirst=True, errors='coerce')
processado_erro['Data EmissÃ£o'] = pd.to_datetime(
    processado_erro['Data EmissÃ£o'], format='mixed', dayfirst=True, errors='coerce')
processado_erro['Data AlteraÃ§Ã£o'] = pd.to_datetime(
    processado_erro['Data AlteraÃ§Ã£o'], format='mixed', dayfirst=True, errors='coerce')

# AtribuiÃ§Ã£o - Fornecedor
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Fornecedor nÃ£o preenchido', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO']] = ['Falta de Fornecedor', 'Campo Fornecedor', 'Dados do Fornecedor']

# AtribuiÃ§Ã£o - de responsÃ¡veis
for row in range(len(processado_erro)):
    for row2 in range(len(info)):
        if info['CAMPO_INFO'][row2] in processado_erro['CAMPO'][row]:
            processado_erro.at[row,
                               'RESPONSÃVEL'] = info.at[row2, 'RESPONSÃVEL_INFO']


# Preenchimento de valores nulos(Status RequisiÃ§Ã£o)
condition = processado_erro['Status Requisicao'].str.contains('-', na=False)
processado_erro.loc[condition, 'Status Requisicao'] = 'OFF LINE'

# Preenchimento de vendas manuais
condition = processado_erro['OBT'].str.contains('MANUAL', na=False)
processado_erro.loc[condition, 'Status Requisicao'] = 'OFF LINE'

# Preenchimento de valores repetidos - falta de informaÃ§Ã£o gerencial
processado_erro.loc[processado_erro['Mensagem Erro'].str.count('nÃ£o preenchid')
                    > 1, 'ORIGEM DO ERRO'] = 'Mais de um campo nÃ£o preenchido'

# RealocaÃ§Ãµes - Suporte KCS (Falta de informaÃ§Ã£o Gerencial e SABRE)
processado_erro.loc[
    (processado_erro['CAMPO'].str.contains('Falta de informaÃ§Ã£o Gerencial')),
    'RESPONSÃVEL'] = 'Suporte KCS'

# RealocaÃ§Ãµes - OperaÃ§Ãµes CORP (Reembolsos Recebidos)
processado_erro.loc[
    (processado_erro['Cliente'] == 'Reembolsos Recebidos'),
    'RESPONSÃVEL'] = 'OperaÃ§Ãµes CORP'

# RealocaÃ§Ãµes - OperaÃ§Ãµes CORP (WS, WT, CT - MANUAL)
processado_erro.loc[
    (processado_erro['OBTS'] == 'MANUAL') &
    (processado_erro['Agente EmissÃ£o'] == 'WS') |
    (processado_erro['Agente EmissÃ£o'] == 'WT') |
    (processado_erro['Agente EmissÃ£o'] == 'CT'),
    'RESPONSÃVEL'] = 'OperaÃ§Ãµes CORP'

# RealocaÃ§Ãµes - OperaÃ§Ãµes CORP (MANUAL e Accounting sem trecho)
processado_erro.loc[
    (processado_erro['OBTS'] == 'MANUAL') &
    (processado_erro['Mensagem Erro'].str.contains('Accounting sem trecho')),
    'RESPONSÃVEL'] = 'OperaÃ§Ãµes CORP'

# RealocaÃ§Ã£o - FCM
clientes_fcm_str = '|'.join(map(str, clientes_fcm['Clientes FCM']))

processado_erro['RESPONSÃVEL'] = processado_erro['Cliente'].apply(
    lambda cliente: 'OperaÃ§Ãµes - FCM' if any(
        cliente_fcm in cliente for cliente_fcm in clientes_fcm['Clientes FCM']) else processado_erro['RESPONSÃVEL'].iloc[0]
)

# RealocaÃ§Ãµes - Suporte KCS (Contrato de fornecedor)
processado_erro.loc[
    (processado_erro['CAMPO'].str.contains('Contrato de fornecedor')),
    'RESPONSÃVEL'] = 'Suporte KCS'

# RealocaÃ§Ãµes - Suporte KCS (TMS - ON LINE)
processado_erro.loc[
    (processado_erro['OBTS'] == 'ARGO(TMS)') &
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['Cliente'].str.contains('Argo It')),
    'RESPONSÃVEL'] = 'Suporte KCS'

# RealocaÃ§Ãµes - OperaÃ§Ãµes CORP (MANUAL)
processado_erro.loc[
    (processado_erro['OBTS'] == 'MANUAL') &
    (processado_erro['RESPONSÃVEL'] == 'Suporte KCS'),
    'RESPONSÃVEL'] = 'OperaÃ§Ãµes CORP'

# RealocaÃ§Ãµes - Suporte KCS (Latam, Gol - Bilhete incompleto)
processado_erro.loc[
    (processado_erro['OBTS'] == 'SABRE') &
    (processado_erro['CAMPO'].str.contains('Bilhete incompleto')) &
    (processado_erro['Fornecedor'].str.contains('Latam|Gol')),
    'RESPONSÃVEL'] = 'Suporte KCS'

# RealocaÃ§Ãµes - Suporte KCS (GOVER)
processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['Status Requisicao'] == 'ON LINE'),
    'RESPONSÃVEL'] = 'Suporte KCS'

# RealocaÃ§Ãµes - Suporte KCS (Falta de Fornecedor - ONLINE - Carro)
processado_erro.loc[
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    (processado_erro['ServiÃ§o'] == 'Carro'),
    'RESPONSÃVEL'] = 'Suporte KCS'

# RealocaÃ§Ãµes - Suporte KCS (Falta de Fornecedor - ONLINE - Lemontech)
processado_erro.loc[
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    (processado_erro['OBTS'] == 'LEMONTECH'),
    'RESPONSÃVEL'] = 'Suporte KCS'

# RealocaÃ§Ãµes - Suporte KCS (Falta de Fornecedor - Offline)
processado_erro.loc[
    (processado_erro['Status Requisicao'] == 'OFF LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')),
    'RESPONSÃVEL'] = 'Suporte KCS'

# RealocaÃ§Ãµes - Suporte KCS (Falta de Fornecedor - ONLINE - Lemontech - Lojas Renner)
processado_erro.loc[
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    (processado_erro['OBTS'] == 'LEMONTECH') &
    (processado_erro['Grupo Empresarial'].str.contains('Grupo Lojas Renner')),
    'RESPONSÃVEL'] = 'Suporte KCS'

# RealocaÃ§Ãµes - Suporte KCS (WS, CT, WT - SABRE)
processado_erro.loc[
    (processado_erro['OBTS'] == 'SABRE') &
    (processado_erro['Agente EmissÃ£o'].str.contains('WS') | processado_erro['Agente EmissÃ£o'].str.contains(
        'CT') | processado_erro['Agente EmissÃ£o'].str.contains('WT')),
    'RESPONSÃVEL'] = 'Suporte KCS'

# RealocaÃ§Ã£o - Suporte Benner (Falta de Fornecedor - Sabre - Hotel - ONLINE)
processado_erro.loc[
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    ((processado_erro['Agente EmissÃ£o'].str.contains('WS')) | (processado_erro['Agente EmissÃ£o'].str.contains('CT'))) &
    ((processado_erro['OBTS'] == 'SABRE') &
     (processado_erro['ServiÃ§o'] == 'Hotel')), ['RESPONSÃVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'SistÃªmico']

# RealocaÃ§Ã£o - OperaÃ§Ãµes CORP (Falta de Fornecedor - KPMG - OFFLINE - Carro)
processado_erro.loc[
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    (processado_erro['Grupo Empresarial'].str.contains('Grupo Kpmg')) &
    (processado_erro['ServiÃ§o'] == 'Carro') &
    (processado_erro['Status Requisicao'] == 'OFF LINE') &
    (pd.to_datetime(processado_erro['Data EmissÃ£o'], dayfirst=True,
     errors='coerce') > pd.to_datetime('06/08/2024', dayfirst=True)),
    ['RESPONSÃVEL', 'CATEGORIA DE ERRO']
] = ['OperaÃ§Ãµes CORP', 'Qualidade dos dados']

# RealocaÃ§Ã£o - Suporte KCS(Gover - Offline - Agente EmissÃ£o - Falta de informaÃ§Ã£o Gerencial)
processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['Status Requisicao'] == 'OFF LINE') &
    (processado_erro['Agente EmissÃ£o'].str.contains('Gover')) &
    (processado_erro['CAMPO'].str.contains('Falta de informaÃ§Ã£o Gerencial')),
    'RESPONSÃVEL'] = 'Suporte KCS'

# RealocaÃ§Ã£o - Suporte KCS(OBTS - Falta de informaÃ§Ã£o Gerencial)
processado_erro.loc[
    (processado_erro['OBTS'].str.contains('ARGO|TMS|GOVER|LEMONTECH')) &
    (processado_erro['CAMPO'].str.contains('Falta de informaÃ§Ã£o Gerencial')),
    'RESPONSÃVEL'] = 'Suporte KCS'

# RealocaÃ§Ã£o - Suporte Benner (Falta de Fornecedor - GOVER - Hotel)
processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['ServiÃ§o'] == 'Hotel') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')),
    ['RESPONSÃVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'SistÃªmico']

# RealocaÃ§Ã£o - Suporte Benner (Falta de Fornecedor - GOVER - Carro - ONLINE)
processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['ServiÃ§o'] == 'Carro') &
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')), [
        'RESPONSÃVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'SistÃªmico']

# RealocaÃ§Ã£o - Suporte Benner (Falta de Fornecedor - Argo - Hotel - ONLINE)
processado_erro.loc[
    (processado_erro['OBTS'] == 'ARGO(TMS)') &
    (processado_erro['ServiÃ§o'] == 'Hotel') &
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')), [
        'RESPONSÃVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'SistÃªmico']

processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['ServiÃ§o'] == 'Hotel') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')), [
        'RESPONSÃVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'SistÃªmico']

# RealocaÃ§Ã£o - Agencia Mercurio York
processado_erro.loc[
    (processado_erro['Cliente'] == 'Agencia Mercurio York'),
    ['OBTS', 'RESPONSÃVEL']] = ['MANUAL', 'OperaÃ§Ãµes Mercurio York']

# RealocaÃ§Ã£o - KONTRIP
processado_erro.loc[
    (processado_erro['OBTS'].str.contains('KONTRIP')),
    'RESPONSÃVEL'] = 'OperaÃ§Ãµes - KONTRIP'

# RealocaÃ§Ã£o - ZUPPER
processado_erro.loc[
    (processado_erro['OBTS'].str.contains('ZUPPER')),
    'RESPONSÃVEL'] = 'OperaÃ§Ãµes - ZUPPER'

# RealocaÃ§Ã£o - Suporte Benner (Cliente FEE no POS)
processado_erro.loc[
    (processado_erro['Cliente Fee POS'] == 'Cliente FEE no POS') &
    (processado_erro['CAMPO'].str.contains('Pagamento nÃ£o permitido para cobranÃ§a')) &
    (processado_erro['Forma Pagamento'] == 'Pagamento direto'),
    ['RESPONSÃVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'SistÃªmico']

# RealocaÃ§Ã£o - Usando planilha de realocaÃ§Ã£o - por localizador
realocacao_handle = processado_erro.merge(
    realocacao[['Handle ACC', 'Equipe ATUAL']], on='Handle ACC', how='left')
processado_erro.loc[realocacao_handle['Equipe ATUAL'].notnull(
), 'RESPONSÃVEL'] = realocacao_handle['Equipe ATUAL']

# RealocaÃ§Ã£o- ConciliaÃ§Ã£o aÃ©rea (Duplicidade de RLOC ou Bilhete duplicado)
processado_erro.loc[
    (processado_erro['ServiÃ§o'].str.contains('AÃ©reo')) &
    ((processado_erro['CAMPO'].str.contains('Duplicidade de RLOC')) |
     (processado_erro['CAMPO'].str.contains('Bilhete duplicado'))),
    'RESPONSÃVEL'] = 'ConciliaÃ§Ã£o aÃ©rea'

# RealocaÃ§Ã£o - Central de EmissÃ£o (Markup) por Agente EmissÃ£o
agentes = [
    'Gessiane Santos Silva', 'Clayton Alves de Rezende', 'Icaro Gabriel Pimentel Gomes Xavier',
    'Carlos Alberto Theodoro da Silva', 'Mylena Mendonca Santos da Silva', 'Wellington Ribeiro da Silva',
    'Itamar de Souza', 'Flavia Constanzi do Nascimento', 'Ana Paula Costa Feitosa', 'Elisete Ferraz de Almeida'
]

processado_erro.loc[
    processado_erro['Agente EmissÃ£o'].isin(
        agentes) | processado_erro['Agente CriaÃ§Ã£o'].isin(agentes),
    ['RESPONSÃVEL', 'CATEGORIA DE ERRO']
] = ['Central de EmissÃ£o', 'Qualidade dos dados']


# RealocaÃ§Ãµes - Suporte KCS (Bilhete UNDEFINED)
processado_erro.loc[
    (processado_erro['OBTS'] == 'SABRE') &
    (processado_erro['Mensagem Erro'].str.contains('UNDEFINED')),
    'RESPONSÃVEL'] = 'Suporte KCS'

# RealocaÃ§Ã£o - ConciliaÃ§Ã£o aÃ©rea (Esta accounting estÃ¡ conciliada no BSP. Bilhete)
processado_erro.loc[
    (processado_erro['Mensagem Erro'].str.contains(
        'Esta accounting estÃ¡ conciliada no BSP. Bilhete:')),
    'RESPONSÃVEL'] = 'ConciliaÃ§Ã£o aÃ©rea'

# RealocaÃ§Ã£o - Central de EmissÃ£o (Markup)
processado_erro['Markup'] = processado_erro['Markup'].astype(float)
processado_erro.loc[processado_erro['Markup']
                    > 0, 'RESPONSÃVEL'] = 'Central de EmissÃ£o'

# RealocaÃ§Ã£o - INOVENTS (Codigo Evento)
processado_erro.loc[
    processado_erro['Canal de Vendas'].str.contains('INOVENTS'),
    'RESPONSÃVEL'] = 'INOVENTS'

# RealocaÃ§Ãµes - EMPRESA - KONTIK BUSINESS TRAVEL
processado_erro.loc[
    processado_erro['RESPONSÃVEL'].str.contains('OperaÃ§Ãµes CORP') | processado_erro['RESPONSÃVEL'].str.contains('OperaÃ§Ãµes - FCM') |
    processado_erro['RESPONSÃVEL'].str.contains('Suporte KCS'),
    'EMPRESA'] = 'KONTIK BUSINESS TRAVEL'

# RealocaÃ§Ãµes - EMPRESA - ZUPPER VIAGENS
processado_erro.loc[
    processado_erro['OBTS'].str.contains('ZUPPER'),
    'EMPRESA'] = 'ZUPPER VIAGENS'

# RealocaÃ§Ãµes - EMPRESA - INOVENTS
processado_erro.loc[
    processado_erro['RESPONSÃVEL'].str.contains('INOVENTS'),
    'EMPRESA'] = 'INOVENTS'

# RealocaÃ§Ãµes - EMPRESA - GRUPO KONTIK
processado_erro.loc[
    processado_erro['RESPONSÃVEL'].str.contains('Central de EmissÃ£o') | processado_erro['RESPONSÃVEL'].str.contains('ConciliaÃ§Ã£o aÃ©rea') |
    processado_erro['RESPONSÃVEL'].str.contains('Suporte Benner'),
    'EMPRESA'] = 'GRUPO KONTIK'

# RealocaÃ§Ãµes - EMPRESA - KONTRIP VIAGENS
processado_erro.loc[
    processado_erro['OBTS'].str.contains('KONTRIP'),
    'EMPRESA'] = 'KONTRIP VIAGENS'

# RealocaÃ§Ãµes - EMPRESA - K-CLUB
processado_erro.loc[
    processado_erro['RESPONSÃVEL'].str.contains('K-CLUB'),
    'EMPRESA'] = 'KTK'

# RealocaÃ§Ã£o - Suporte Benner (NÃ£o foi possÃ­vel definir o Local de destino! - Suporte Benner)
processado_erro.loc[
    (processado_erro['Mensagem Erro'].str.contains(
        'NÃ£o foi possÃ­vel definir o Local de destino!')),
    'RESPONSÃVEL'] = 'Suporte Benner'

# VerificaÃ§Ã£o de duplicidade(Gover)
cont_localizador = processado_erro['Localizadora'].tolist()

for row in range(len(processado_erro)):
    if cont_localizador.count(processado_erro['Localizadora'][row]) > 10 and \
            'GOVER' in processado_erro['OBTS'][row]:
        processado_erro.at[row, 'CAMPO'] = 'Falha no processo de integraÃ§Ã£o'
        processado_erro.at[row,
                           'ORIGEM DO ERRO'] = 'Vendas duplicadas'
        processado_erro.at[row, 'TIPO DE ERRO'] = 'Sistema'
        processado_erro.at[row, 'RESPONSÃVEL'] = 'Suporte Benner'
        processado_erro.at[row, 'CATEGORIA DE ERRO'] = 'SistÃªmico'

        if 'Vendas duplicadas' in str(processado_erro.at[row, 'ORIGEM DO ERRO']):
            print(f'\033[1;31m-Verifique o localizador "{processado_erro.at[row, "Localizadora"]}"'
                  f' e requisiÃ§Ã£o "{processado_erro.at[row, "RequisiÃ§Ã£o"]}",'
                  f' feitas pelo consultor "{processado_erro.at[row, "Agente EmissÃ£o"]}"\033[m')
            break

# Salvar o arquivo original
processado_erro.to_excel(data_path + 'Processado Erro.xlsx', index=False)

# SeleÃ§Ã£o de colunas
colunas = {
    'A': 'Handle PNR', 'B': 'Handle ACC', 'C': 'Sequencia', 'D': 'Data InclusÃ£o',
    'E': 'Data AlteraÃ§Ã£o', 'F': 'MÃªs AlteraÃ§Ã£o', 'G': 'Dias Parados no Erro', 'H': 'Localizadora', 'I': 'OBT', 'J': 'Pax',
    'K': 'Agente EmissÃ£o', 'L': 'Agente CriaÃ§Ã£o', 'M': 'Data EmissÃ£o', 'N': 'RequisiÃ§Ã£o', 'O': 'Local Retirada',
    'P': 'Status Requisicao', 'Q': 'Forma Pagamento', 'R': 'Forma Recebimento',
    'S': 'ServiÃ§o', 'T': 'Cancelado', 'U': 'Grupo Empresarial', 'V': 'Cliente',
    'W': 'Cliente Fee POS', 'X': 'Fornecedor', 'Y': 'Bilhete', 'Z': 'Canal de Vendas',
    'AA': 'Codigo Evento', 'AB': 'Tarifa', 'AC': 'Taxa', 'AD': 'Outras Taxas',
    'AE': 'Taxa DU', 'AF': 'Taxa BR', 'AG': 'Taxa Extra', 'AH': 'Fee',
    'AI': 'ObservaÃ§Ã£o', 'AJ': 'Mensagem Erro', 'AK': 'OBTS', 'AL': 'CAMPO',
    'AM': 'ORIGEM DO ERRO', 'AN': 'TIPO DE ERRO', 'AO': 'CATEGORIA DE ERRO',
    'AP': 'EMPRESA', 'AQ': 'RESPONSÃVEL'
}

# InserÃ§Ã£o de colunas
for col, nome in colunas.items():
    relatorio_base[col + '1'] = nome

# InserÃ§Ã£o de valores
for row in range(len(processado_erro)):
    for col, nome in colunas.items():
        relatorio_base[col + str(row + 2)] = processado_erro[nome][row]

# PersonalizaÃ§Ã£o de cÃ©lulas


def personalizacao(relatorio):
    # Definindo cores para o cabeÃ§alho
    colunas_padrao = PatternFill(
        start_color="404654", end_color="404654", fill_type="solid")  # CINZA
    colunas_adicionais = PatternFill(
        start_color="c4d72f", end_color="c4d72f", fill_type="solid")  # VERDE
    cor_da_fonte = Font(color="FFFFFF", bold=True)  # BRANCO

    # Aplicar a formataÃ§Ã£o aos cabeÃ§alhos
    for coluna in range(1, 44):  # Colunas de A atÃ© AS
        letra_coluna = relatorio.cell(row=1, column=coluna).column_letter
        cabecalho = relatorio[letra_coluna + '1']
        cabecalho.fill = colunas_padrao
        cabecalho.font = cor_da_fonte
        if letra_coluna in ['F', 'G', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ']:
            cabecalho.fill = colunas_adicionais

    # Definindo cores para aging de dias
    for row in relatorio.iter_rows(min_row=2, min_col=7, max_col=7):
        cell_value = row[0].value
        if cell_value in ['16 a 23 dias', '24 a 31 dias', '31 dias ou +']:
            row[0].font = Font(color="FF0000")


personalizacao(relatorio_base)
# Salvar relatorio base
wb.save(data_path + 'Relatorio - Dash.xlsx')

# Carregar a planilha original
relatorio_dash = pd.read_excel(
    data_path + 'Relatorio - Dash.xlsx', sheet_name='Processado Erro - BASE')
data_path = r'PROCESSADO ERRO\Analise de Dados\\'

empresas = relatorio_dash['EMPRESA'].unique()

# Defina o caminho do diretÃ³rio onde os arquivos serÃ£o salvos
output_folder = data_path + 'EMPRESAS'

# Verifique se o diretÃ³rio existe; se nÃ£o, crie-o
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

for empresa in empresas:
    # Criar uma nova planilha para cada empresa
    df_empresa = relatorio_dash[relatorio_dash['EMPRESA'] == empresa]
    output_file = os.path.join(output_folder, f'Relatorio - {empresa}.xlsx')

    # Salvando o DataFrame da empresa em um arquivo Excel
    df_empresa.to_excel(output_file, index=False)

    # Abrindo o arquivo Excel para aplicar formataÃ§Ã£o
    wb = load_workbook(output_file)
    ws = wb.active  # Pegando a primeira aba

    # Aplicando a personalizaÃ§Ã£o
    personalizacao(ws)

    # Salvando as mudanÃ§as no arquivo Excel
    wb.save(output_file)

    print(f'Relatorio - {empresa} - Salvo com sucesso!')
print()
print('\033[1;32m-RelatÃ³rio gerado com sucesso!\033[m')
