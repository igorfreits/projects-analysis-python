import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

caminho_arquivo = 'data-analysis-python/Controle - Data Quality - FCM/'
arquivo_fcm = caminho_arquivo + 'Controle - FCM.xlsx'

controle_xls = pd.ExcelFile(arquivo_fcm)
before = controle_xls.parse(sheet_name='Before')
after = controle_xls.parse(sheet_name='After')

# Garante que as chaves estejam como texto
for col_key in ['Handle ACC', 'Grupo Empresarial', 'Cliente']:
    before[col_key] = before[col_key].astype(str)
    after[col_key] = after[col_key].astype(str)

# Definir chaves fixas para merge
colunas_fixes = ['Handle ACC', 'Grupo Empresarial', 'Cliente', 'Data de Emissão', 'Produto']

# Faz merge usando as chaves
merged = pd.merge(before, after, on=colunas_fixes, how='inner', suffixes=('_before', '_after'))

# Lista de colunas a comparar (excluindo as chaves)
colunas_comparacao = [col for col in before.columns if col not in colunas_fixes]

diferencas = []
for col in colunas_comparacao:
    col_b = f"{col}_before"
    col_a = f"{col}_after"

    # Limpeza e padronização de campos vazios
    merged[col_b] = merged[col_b].astype(str).str.strip().replace({'': pd.NA, 'nan': pd.NA})
    merged[col_a] = merged[col_a].astype(str).str.strip().replace({'': pd.NA, 'nan': pd.NA})

    # Apenas onde os valores são diferentes e não estão ambos vazios
    diferentes = merged[
        (merged[col_b] != merged[col_a]) &
        ~(merged[col_b].isna() & merged[col_a].isna())
    ][colunas_fixes + [col_b, col_a]]

    diferentes['Campo'] = col
    diferentes = diferentes.rename(columns={col_b: 'Valor Antes', col_a: 'Valor Depois'})

    colunas_ordenadas = colunas_fixes + ['Campo', 'Valor Antes', 'Valor Depois']
    diferentes = diferentes[colunas_ordenadas]

    diferencas.append(diferentes)


resultado_final = pd.concat(diferencas, ignore_index=True)

with pd.ExcelWriter(arquivo_fcm, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    resultado_final.to_excel(writer, sheet_name='Comparativo Benner', index=False)

# Formatação
wb = openpyxl.load_workbook(arquivo_fcm)
ws = wb['Comparativo Benner']

# Congela o cabeçalho
ws.freeze_panes = 'A2'

# Estilo
color_cabecalho = PatternFill(start_color='404654', end_color='404654', fill_type='solid')
fonte_cabecalho = Font(color='C4D72F', bold=True)

# Aplica estilo nas células do cabeçalho
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = color_cabecalho
    cell.font = fonte_cabecalho

# Ajuste automático de largura de coluna (opcional)
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save(arquivo_fcm)

print("Comparação concluída com sucesso.")
