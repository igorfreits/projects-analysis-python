import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill,Alignment
from openpyxl import load_workbook
import os
import numpy as np
import re
import shutil
import win32com.client


# Função para converter .xls para .xlsx
def converter_xls_para_xlsx(caminho_arquivo):
    if not caminho_arquivo or not isinstance(caminho_arquivo, str):
        raise ValueError("Caminho do arquivo inválido.")

    if not caminho_arquivo.lower().endswith(".xls"):
        raise ValueError("O arquivo fornecido não é um .xls")

    if not os.path.exists(caminho_arquivo):
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_arquivo}")

    base, _ = os.path.splitext(caminho_arquivo)
    novo_caminho = base + ".xlsx"

    try:
        df = pd.read_excel(caminho_arquivo, engine="xlrd")
        if df.empty:
            raise ValueError("O arquivo está vazio.")

        df.to_excel(novo_caminho, index=False, engine="openpyxl")
        return novo_caminho
    except Exception as e:
        print(f"Erro ao converter {caminho_arquivo}: {e}")
        return None


# ---Caminho dos arquivos---
data_path = 'C:\\Users\\' + os.getlogin() + '\\Desktop\\DOCS\\data-analysis-python\\Integration-Quality\\'
os.makedirs(data_path, exist_ok=True)

# Inicializa o Outlook
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

# Acessa a caixa de e-mail compartilhada
caixa_compartilhada = outlook.Folders.Item("Suporte Benner")

# Acessa a pasta "Portal Wes - ERRO"
pasta = caixa_compartilhada.Folders.Item("Caixa de Entrada").Folders.Item("Portal Wes - ERRO")

# ---Captura do e-mail e salvamento do anexo---
# Filtra os e-mails recebidos hoje na pasta "Portal Wes - ERRO", pega o mais recente
# e salva o primeiro anexo no diretório de trabalho para processamento.
# Captura o email do dia atual
hoje = datetime.now().date()
emails = pasta.Items
emails.Sort("[ReceivedTime]", True)  # Ordena por data decrescente para pegar o mais recente
email_do_dia = [email for email in emails if email.ReceivedTime.date() == hoje]

if not email_do_dia:
    raise RuntimeError(f"Nenhum e-mail encontrado na pasta 'Portal Wes - ERRO' para hoje ({hoje.strftime('%d/%m/%Y')}).")

# Pega o e-mail mais recente do dia e salva o anexo
usuario = os.getlogin()
ultimo_email = email_do_dia[0]
caminho_anexo = None
if ultimo_email.Attachments.Count > 0:
    anexo = ultimo_email.Attachments.Item(1)
    nome_anexo = anexo.FileName
    caminho_anexo = os.path.join(data_path, nome_anexo)

    arquivo_ja_baixado = (
        os.path.exists(caminho_anexo) and
        datetime.fromtimestamp(os.path.getmtime(caminho_anexo)).date() == hoje
    )
    if arquivo_ja_baixado:
        print(f"Anexo '{nome_anexo}' já baixado hoje, pulando download.")
    else:
        anexo.SaveAsFile(caminho_anexo)
        print(f"Anexo '{nome_anexo}' salvo em: {caminho_anexo}")
else:
    raise RuntimeError(f"O e-mail de hoje ({hoje.strftime('%d/%m/%Y')}) não possui anexo.")

arquivo_xls = caminho_anexo

# Verifica se o arquivo .xls existe e converte antes de ler
if os.path.exists(arquivo_xls):
    arquivo_xls = converter_xls_para_xlsx(arquivo_xls)

# Arquivos Excel
try:
    processado_erro = pd.read_excel(arquivo_xls)
except Exception as e:
    raise RuntimeError(f"Erro ao carregar arquivo de erros: {e}")

realocacao = pd.read_excel(data_path + 'Realocacao.xlsx')
 
# Declaração de guias - Relatório DASH
dash = load_workbook(data_path + 'Relatorio - Dash.xlsx')
relatorio_base = dash['Processado Erro - BASE']
 
# Limpeza de planilha
relatorio_base.delete_rows(2, relatorio_base.max_row)
relatorio_base.delete_cols(2, relatorio_base.max_column)
 
# Criação de colunas
processado_erro['Aging Alteração'] = ''
processado_erro['Aging Inclusão'] = ''
processado_erro['OBTS'] = ''
processado_erro['CAMPO'] = 'Não identificado'
processado_erro['ORIGEM DO ERRO'] = 'Análise Benner'
processado_erro['TIPO DE ERRO'] = 'Sistema'
processado_erro['CATEGORIA DE ERRO'] = 'Qualidade dos dados'
processado_erro['EMPRESA'] = ''
processado_erro['RESPONSÁVEL'] = 'Operações - CORP'
 
# Preenchimento de valores nulos
processado_erro['Mensagem Erro'] = processado_erro['Mensagem Erro'].fillna('Erro não localizado')
processado_erro = processado_erro.fillna('-')
  
# Tratamento da coluna 'OBT'
processado_erro['OBTS'] = processado_erro['OBT']
processado_erro['OBTS'] = processado_erro['OBTS'].str.replace(
    'TMS', 'ARGO(TMS)')
 
# Aplicação da máscara para substituir valores na coluna 'OBTS' - ZUPPER
mask_zupper = (processado_erro['Canal de Vendas'].str.contains('ZUPPER', case=False) |
               processado_erro['Grupo Empresarial'].str.contains('Grupo Zupper', case=False))
processado_erro['OBTS'] = processado_erro['OBTS'].mask(mask_zupper, 'ZUPPER')
 
# Aplicação da máscara para substituir valores na coluna 'OBTS' - KONTRIP
mask_kontrip = processado_erro['Canal de Vendas'].str.contains(
    'KONTRIP', case=False)
processado_erro['OBTS'] = processado_erro['OBTS'].mask(mask_kontrip, 'KONTRIP')
 
# Tratamento de colunas - formatação de strings
processado_erro['Mensagem Erro'] = processado_erro['Mensagem Erro'].astype(str)

#---Formatação de datas e aging---
processado_erro['Aging Inclusão'] = (
    datetime.now() - pd.to_datetime(processado_erro['Data Inclusão'].str[:10], format='%d/%m/%Y', errors='coerce')).dt.days
processado_erro['Aging Alteração'] = (
    datetime.now() - pd.to_datetime(processado_erro['Data Alteração'].str[:10], format='%d/%m/%Y', errors='coerce')).dt.days

# Leitura de data de alteração - Edição não permitida
processado_erro.loc[processado_erro['TIPO DE ERRO'].str.contains('Edição não Permitida'), 'Aging Inclusão'] = (
    datetime.now() - pd.to_datetime(processado_erro['Data Alteração'].str[:10], format='%d/%m/%Y', errors='coerce')
).dt.days

# Leitura de data de alteração - Bilhete duplicado
processado_erro.loc[processado_erro['CAMPO'].str.contains('Bilhete duplicado'), 'Aging Inclusão'] = (
    datetime.now() - pd.to_datetime(processado_erro['Data Alteração'].str[:10], format='%d/%m/%Y', errors='coerce')
).dt.days

# Preenchimento de data de emissão, se estiver vazia
processado_erro.loc[processado_erro['Data Emissão'].str.contains(
    '-'), 'Data Emissão'] = processado_erro['Data Inclusão']
 
# Categorização de dias parados
limites = [0, 3, 6, 9, 16, 24, 31, float('inf')]
rotulos = ['0 a 02 dias', '03 a 05 dias', '06 a 08 dias', '09 a 15 dias', '16 a 23 dias', '24 a 31 dias', '31 dias ou +']
 
# Aplicar categorização de dias parados
processado_erro['Aging Inclusão'] = pd.cut(processado_erro['Aging Inclusão'], bins=limites, labels=rotulos, right=False, include_lowest=True)
processado_erro['Aging Alteração'] = pd.cut(processado_erro['Aging Alteração'], bins=limites, labels=rotulos, right=False, include_lowest=True)

# Aplicar formatação com dayfirst=True para ajustar o formato de data "DD/MM/AAAA HH:MM"
processado_erro['Data Inclusão'] = pd.to_datetime(processado_erro['Data Inclusão'], format='mixed', dayfirst=True, errors='coerce')
processado_erro['Data Emissão'] = pd.to_datetime(processado_erro['Data Emissão'], format='mixed', dayfirst=True, errors='coerce')
processado_erro['Data Alteração'] = pd.to_datetime(processado_erro['Data Alteração'], format='mixed', dayfirst=True, errors='coerce')

# Tratamento de Status Requisição
processado_erro['Status Requisicao'] = np.where(processado_erro['Status Requisicao'].str.contains('-', na=False), 'OFF LINE', processado_erro['Status Requisicao'])
processado_erro['Status Requisicao'] = np.where(processado_erro['OBT'].str.contains('MANUAL', na=False), 'OFF LINE', processado_erro['Status Requisicao'])


#---Tratamento de Erros - Padrões---

# Preenchimento de valores repetidos - falta de informação gerencial
multiplos_campos = processado_erro['Mensagem Erro'].str.lower().str.count('não preenchid') > 1
processado_erro.loc[multiplos_campos, ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Mais de um campo não preenchido','Falta de informação Gerencial','Dados do Fornecedor','Qualidade dos dados']

# Para um único campo não preenchido – extrai o nome do campo
unico_campo = processado_erro['Mensagem Erro'].str.lower().str.count('não preenchid') == 1

# Extrai o nome do campo antes de "não preenchido" e remove espaços
campo_extraido = processado_erro.loc[unico_campo, 'Mensagem Erro'].str.extract(
    r'^(.*?)\s*não preenchid[ao]', expand=False, flags=re.IGNORECASE
).str.strip()

# Cria o DataFrame com os valores fixos + o campo extraído
valores = pd.DataFrame({
    'CAMPO': campo_extraido,
    'ORIGEM DO ERRO': 'Falta de informação Gerencial',
    'TIPO DE ERRO': 'Dados do Fornecedor',
    'CATEGORIA DE ERRO': 'Qualidade dos dados'
}, index=campo_extraido.index)

processado_erro.loc[unico_campo, ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = valores

# ── Bloco 1: Duplicidade de Contabilização / Qualidade dos dados ──────────────

# PNR duplicado
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Pnr já existente', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Duplicidade de RLOC', 'Campo RLOC', 'Duplicidade de Contabilização', 'Qualidade dos dados']

# Duplicidade de Bilhete
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Verificação de bilhetes: Bilhete', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Bilhete duplicado', 'Bilhete Já Contabilizado', 'Duplicidade de Contabilização', 'Qualidade dos dados']

# ── Bloco 2: Dados do Fornecedor / Qualidade dos dados ────────────────────────

# Falta de Fornecedor
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Fornecedor não preenchido!', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Falta de Fornecedor', 'Campo Fornecedor', 'Dados do Fornecedor', 'Qualidade dos dados']

# Falta de status no trecho
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Falta informar o status no trecho', case=False) |
                    processado_erro['Mensagem Erro'].str.contains('Accouting aérea não possui trecho', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Falta informação nos trechos', 'Atualização de Trechos', 'Dados do Fornecedor', 'Qualidade dos dados']

# Accounting sem trecho (standalone)
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Accounting sem trecho', case=False) &
                    ~processado_erro['Mensagem Erro'].str.contains('Falta informar o status no trecho', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Falta informação nos trechos', 'Atualização de Trechos', 'Dados do Fornecedor', 'Qualidade dos dados']

# Sem permissão para tipo de pagamento/recebimento
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Este cliente não possui permissão para usar este tipo de pagamento', case=False) |
                    processado_erro['Mensagem Erro'].str.contains('-> Configurações do sistema turismo', case=False) |
                    processado_erro['Mensagem Erro'].str.contains('Módulo Operações ', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Pagamento não permitido para cobrança', 'Forma PG. e REC.', 'Dados do Fornecedor', 'Qualidade dos dados']

# Contrato de fornecedor/cliente não encontrado
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Não foi possível encontrar um contrato válido para o fornecedor', case=False) |
                    processado_erro['Mensagem Erro'].str.contains('Necessário cadastrar um contrato para o cliente', case=False) |
                    processado_erro['Mensagem Erro'].str.contains('Consolidador é obrigatório para este fornecedor', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Contrato de fornecedor', 'Análise Cadastro', 'Dados do Fornecedor', 'Qualidade dos dados']

# Contrato vencido/vigência encerrada
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('vig[eê]ncia|contrato vencido|contrato expirado', case=False, regex=True),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Contrato vencido', 'Análise Cadastro', 'Dados do Fornecedor', 'Qualidade dos dados']

# Documento inválido (CPF/CNPJ)
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('CPF inv[aá]lid|CNPJ inv[aá]lid|documento inv[aá]lid', case=False, regex=True),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Documento inválido', 'Cadastro Pax', 'Dados do Fornecedor', 'Qualidade dos dados']

# Limite de crédito excedido
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('limite de cr[eé]dito|saldo insuficiente|cr[eé]dito insuficiente', case=False, regex=True),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Limite de crédito excedido', 'Financeiro', 'Dados do Fornecedor', 'Qualidade dos dados']

# Soma das tarifas dos registros múltiplos
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('A soma das outras taxas dos registros múltiplos é menor que a outras taxas da accounting', case=False) |
                    processado_erro['Mensagem Erro'].str.contains('A soma das tarifas dos registros múltiplos é menor que a tarifa da accounting', case=False) |
                    processado_erro['Mensagem Erro'].str.contains('A soma das tarifas dos registros múltiplos ultrapassa a tarifa da accounting', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Soma das tarifas dos registros múltiplos', 'Validação de Tarifas', 'Dados do Fornecedor', 'Qualidade dos dados']

# Enriquecimento de dados pendente
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Enriquecimento de dados pendente', case=False, regex=True),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Enriquecimento de dados pendente', 'Validação de Dados', 'Dados do Fornecedor', 'Qualidade dos dados']

# Não foi possível encontrar configuração válida para Administradora
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Não foi possível encontrar configuração válida para Administradora', case=False, regex=True),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Configuração de Administradora', 'Validação de Dados', 'Dados do Fornecedor', 'Qualidade dos dados']

# Venda sem bilhete
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Necessário informar o bilhete', case=False, regex=True),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Venda sem bilhete', 'Validação de Dados', 'Dados do Fornecedor', 'Qualidade dos dados']

# ── Bloco 3: Dados Gerenciais / Qualidade dos dados ───────────────────────────

# Validação do documento: período financeiro fechado
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Documento está dentro de um período financeiro já fechado!', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Documento fora do período financeiro', 'Financeiro Conciliado', 'Dados Gerenciais', 'Qualidade dos dados']

# Rateio de centro de custo/projeto
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('centro de custo/projeto', case=False) |
                    processado_erro['Mensagem Erro'].str.contains('Ocorreu a seguinte exceção ao inserir o item da ordem de venda', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Falta de informação Gerencial', 'Rateio de centro de custo/projeto', 'Dados Gerenciais', 'Qualidade dos dados']

# Caractere especial em campo Gerencial
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('A name contained an invalid character', case=False) |
                    processado_erro['Mensagem Erro'].str.contains('</', case=False) |
                    processado_erro['Mensagem Erro'].str.contains('Whitespace is not allowed at this location', case=False) |
                    processado_erro['Mensagem Erro'].str.contains('Erro ao processar PNR  | A semi colon character was expected', case=False) |
                    processado_erro['Mensagem Erro'].str.contains("'*' is not a valid integer value", case=False) |
                    processado_erro['Mensagem Erro'].str.contains('Ol', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Caractere inválido', 'Caractere invalido em campo gerencial', 'Dados Gerenciais', 'Qualidade dos dados']

# ── Bloco 4: Edição não Permitida / Processo Operacional ──────────────────────

# Conciliado BSP
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Esta accounting está conciliada no BSP. Bilhete', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Bilhete conciliado', 'Concilização eletrônica Aérea', 'Edição não Permitida', 'Processo Operacional']

# Controle de Comissão pós paga
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('A mesma está ligada ao controle de comissão pós paga', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Controle de comissão pós paga', 'Financeiro Conciliado', 'Edição não Permitida', 'Processo Operacional']

# Conciliação de cartão
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('A mesma está ligada a conciliação de cartão', case=False) |
                    processado_erro['Mensagem Erro'].str.contains('possui uma transação de cartão efetivada', case=False) |
                    processado_erro['Mensagem Erro'].str.contains('Postar Venda Não foi possível identificar o vínculo ', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Conciliação de Cartão', 'Financeiro Conciliado', 'Edição não Permitida', 'Processo Operacional']

# ── Bloco 5: Sistema / Sistêmico ──────────────────────────────────────────────

# Cadastro enviado errado no Benner - Zupper
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Não foi possível definir o Local de destino!', case=False),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Cadastro enviado errado no Benner', 'Cadastro incompleto', 'Sistema', 'Sistêmico']

# Falha de comunicação/timeout
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('timeout|time out|connection|falha de comunica', case=False, regex=True) |
                        processado_erro['Mensagem Erro'].str.contains('Ocorrência no método TBSistemaWrapper', case=False, regex=True),
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO']] = ['Falha de comunicação', 'Falha de Integração', 'Sistema', 'Sistêmico']



#---Realocações - Responsáveis,Categorias de Erro, Origens do Erro e Tipo de Erro---
# Realocações - Suporte KCS (Falta de informação Gerencial e SABRE)
processado_erro.loc[
    (processado_erro['CAMPO'].str.contains('Falta de informação Gerencial')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocações - Operações - CORP (Reembolsos Recebidos)
processado_erro.loc[
    (processado_erro['Cliente'] == 'Reembolsos Recebidos'),
    'RESPONSÁVEL'] = 'Operações - CORP'
 
# Realocações - Operações - CORP (OFFLINE - duplicado - WS, WT, CT - MANUAL)
processado_erro.loc[
    (processado_erro['OBTS'] == 'MANUAL') &
    (processado_erro['Agente Emissão'].isin(['WS', 'WT', 'CT'])),
    'RESPONSÁVEL'] = 'Operações - CORP'
 
# Realocações - Operações - CORP (MANUAL e Accounting sem trecho)
processado_erro.loc[
    (processado_erro['OBTS'] == 'MANUAL') &
    (processado_erro['Mensagem Erro'].str.contains('Accounting sem trecho')),
    'RESPONSÁVEL'] = 'Operações - CORP'
 
# Realocação - FCM
# clientes_fcm_str = '|'.join(map(str, clientes_fcm['Clientes FCM']))
 
# processado_erro['RESPONSÁVEL'] = processado_erro['Cliente'].apply(
#     lambda cliente: 'Operações - FCM' if any(cliente_fcm in cliente for cliente_fcm in clientes_fcm['Clientes FCM']) else processado_erro['RESPONSÁVEL'].iloc[0]
# )

# Realocação - Operações - FCM
processado_erro.loc[
    (processado_erro['Canal de Vendas'].str.contains('FCM', case=False)),
    'RESPONSÁVEL'] = 'Operações - FCM'

# Realocações - Suporte KCS (Contrato de fornecedor)
processado_erro.loc[
    (processado_erro['CAMPO'].str.contains('Contrato de fornecedor')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocações - Suporte KCS (TMS - ON LINE)
processado_erro.loc[
    (processado_erro['OBTS'] == 'ARGO(TMS)') &
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['Cliente'].str.contains('Argo It')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocações - Operações - CORP (MANUAL)
processado_erro.loc[
    (processado_erro['OBTS'] == 'MANUAL') &
    (processado_erro['RESPONSÁVEL'] == 'Suporte KCS'),
    'RESPONSÁVEL'] = 'Operações - CORP'
 
# Realocações - Suporte KCS (Latam, Gol - Bilhete incompleto)
processado_erro.loc[
    (processado_erro['OBTS'] == 'SABRE') &
    (processado_erro['CAMPO'].str.contains('Bilhete incompleto')) &
    (processado_erro['Fornecedor'].str.contains('Latam|Gol')),
    'RESPONSÁVEL'] = 'Suporte KCS'
   
# Realocações - Suporte KCS (GOVER)
processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['Status Requisicao'] == 'ON LINE'),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocações - Suporte KCS (Falta de Fornecedor - ONLINE - Carro)
processado_erro.loc[
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    (processado_erro['Serviço'] == 'Carro'),
    'RESPONSÁVEL'] = 'Suporte KCS'

# Realocações - Suporte KCS (Falta de Fornecedor - ONLINE - Lemontech)
processado_erro.loc[
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    (processado_erro['OBTS'] == 'LEMONTECH'),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocações - Suporte KCS (Falta de Fornecedor - Offline)
processado_erro.loc[
    (processado_erro['Status Requisicao'] == 'OFF LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocações - Suporte KCS (Falta de Fornecedor - ONLINE - Lemontech - Lojas Renner)
processado_erro.loc[
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    (processado_erro['OBTS'] == 'LEMONTECH') &
    (processado_erro['Grupo Empresarial'].str.contains('Grupo Lojas Renner')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocações - Suporte KCS (ONLINE - WS, CT, WT - SABRE)
processado_erro.loc[
    (processado_erro['OBTS'] == 'SABRE') &
    (processado_erro['Agente Emissão'].str.contains('WS') | processado_erro['Agente Emissão'].str.contains('CT') | processado_erro['Agente Emissão'].str.contains('WT')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocação - Suporte Benner (Falta de Fornecedor - Sabre - Hotel - ONLINE)
# Ajuste de responsável e categoria de erro
processado_erro.loc[
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    ((processado_erro['Agente Emissão'].str.contains('WS')) | (processado_erro['Agente Emissão'].str.contains('CT'))) &
    ((processado_erro['OBTS'] == 'SABRE') &
    (processado_erro['Serviço'] == 'Hotel')),['RESPONSÁVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'Sistêmico']

# Realocação - Operações - CORP (Falta de Fornecedor - KPMG - OFFLINE - Carro - Data Emissão > 06/08/2024)
# Ajuste de responsável e categoria de erro
processado_erro.loc[
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    (processado_erro['Grupo Empresarial'].str.contains('Grupo Kpmg')) &
    (processado_erro['Serviço'] == 'Carro') &
    (processado_erro['Status Requisicao'] == 'OFF LINE') &
    (pd.to_datetime(processado_erro['Data Emissão'], dayfirst=True, errors='coerce') > pd.to_datetime('06/08/2024', dayfirst=True)),
    ['RESPONSÁVEL', 'CATEGORIA DE ERRO']
    ] = ['Operações - CORP', 'Qualidade dos dados']
 
# Realocação - Suporte KCS(Gover - Offline - Agente Emissão - Falta de informação Gerencial)
processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['Status Requisicao'] == 'OFF LINE') &
    (processado_erro['Agente Emissão'].str.contains('Gover')) &
    (processado_erro['CAMPO'].str.contains('Falta de informação Gerencial')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocação - Suporte KCS(OBTS - Falta de informação Gerencial)
processado_erro.loc[
    (processado_erro['OBTS'].str.contains('ARGO|TMS|GOVER|LEMONTECH')) &
    (processado_erro['CAMPO'].str.contains('Falta de informação Gerencial')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocação - Suporte Benner (Falta de Fornecedor - GOVER - Hotel)
# Ajuste de responsável e categoria de erro
processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['Serviço'] == 'Hotel') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')),
    ['RESPONSÁVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'Sistêmico']

# Realocação - Suporte Benner (Falta de Fornecedor - GOVER - Carro - ONLINE)
# Ajuste de responsável e categoria de erro
processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['Serviço'] == 'Carro') &
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')),['RESPONSÁVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'Sistêmico']
 
# Realocação - Suporte Benner (Falta de Fornecedor - Argo - Hotel - ONLINE)
# Ajuste de responsável e categoria de erro
processado_erro.loc[
    (processado_erro['OBTS'] == 'ARGO(TMS)') &
    (processado_erro['Serviço'] == 'Hotel') &
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')),['RESPONSÁVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'Sistêmico']

# Realocação - Suporte Benner (Falta de Fornecedor - GOVER - Hotel)
# Ajuste de responsável e categoria de erro
processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['Serviço'] == 'Hotel') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')),['RESPONSÁVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'Sistêmico']

# Realocação - Operações  - Mercurio York
processado_erro.loc[
    (processado_erro['Cliente'] == 'Agencia Mercurio York'),
    ['OBTS', 'RESPONSÁVEL']] = ['MANUAL', 'Operações - Mercurio York']
 
# Realocação - Suporte Benner (Cliente FEE no POS)
# Ajuste de responsável e categoria de erro
processado_erro.loc[
    (processado_erro['Cliente Fee POS'] == 'Cliente FEE no POS') &
    (processado_erro['CAMPO'].str.contains('Pagamento não permitido para cobrança')) &
    (processado_erro['Forma Pagamento'] == 'Pagamento direto'),
    ['RESPONSÁVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'Sistêmico']
 
# Realocação - Usando planilha de realocação - por localizador
realocacao_handle = processado_erro.merge(realocacao[['Handle ACC', 'Equipe ATUAL']], on='Handle ACC', how='left')
processado_erro.loc[realocacao_handle['Equipe ATUAL'].notnull(), 'RESPONSÁVEL'] = realocacao_handle['Equipe ATUAL']
 
# Realocação- Conciliação aérea (Duplicidade de RLOC ou Bilhete duplicado)
processado_erro.loc[
    (processado_erro['Serviço'].str.contains('Aéreo')) &
    ((processado_erro['CAMPO'].str.contains('Duplicidade de RLOC')) |
     (processado_erro['CAMPO'].str.contains('Bilhete duplicado'))),
    'RESPONSÁVEL'] = 'Conciliação aérea'

# Realocação - Central de Emissão (Agente Emissão ou Agente Criação)
agentes = ['Alexander Perez Alves','Ana Paula Costa Feitosa','Beatrys Ferreira Rocha','Clayton Alves de Rezende','Flavia Constanzi do Nascimento',
           'Gisele Soares Carmo','Icaro Gabriel Pimentel Gomes Xavier','Itamar de Souza','Juliana dos Santos Pinto','Leandra Vitoria Gomes Santos',
           'MARKUP','Mylena Mendonca Santos da Silva','RAFAEL LIMA DA SILVA','teste_produtosaéreos','Wagney Araujo de Oliveira','Wellington Ribeiro da Silva'
]

# Realocação - Central de Emissão (Agente Emissão ou Agente Criação)
# Ajuste de responsável e categoria de erro
processado_erro.loc[
    processado_erro['Agente Emissão'].isin(agentes) | processado_erro['Agente Criação'].isin(agentes),
    ['RESPONSÁVEL', 'CATEGORIA DE ERRO']
] = ['Central de Emissão', 'Qualidade dos dados']


# Realocações - Suporte KCS (Bilhete UNDEFINED)
processado_erro.loc[
    (processado_erro['OBTS'] == 'SABRE') &
    (processado_erro['Mensagem Erro'].str.contains('UNDEFINED')),
    'RESPONSÁVEL'] = 'Suporte KCS'

# Realocações - Suporte KCS (The INSERT statement conflicted with the FOREIGN KEY)
# Ajuste de Categoria de erro, Campo, Origem do Erro e Tipo de Erro
processado_erro.loc[
    (processado_erro['Mensagem Erro'].str.contains(
        'The INSERT statement conflicted with the FOREIGN KEY')),
    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO', 'RESPONSÁVEL']
] = ['Falta de informação Gerencial', 'Mais de um campo não preenchido', 'Dados do Fornecedor', 'Sistêmico', 'Suporte KCS']
 
# Realocação - Conciliação aérea (Esta accounting está conciliada no BSP. Bilhete)
processado_erro.loc[
    (processado_erro['Mensagem Erro'].str.contains(
        'Esta accounting está conciliada no BSP. Bilhete:')),
    'RESPONSÁVEL'] = 'Conciliação aérea'
 
# Realocação - Central de Emissão (Markup)
processado_erro['Markup'] = pd.to_numeric(processado_erro['Markup'], errors='coerce').fillna(0)
processado_erro.loc[processado_erro['Markup']
                    > 0, 'RESPONSÁVEL'] = 'Central de Emissão'

# Realocação - Central de Emissão (Enriquecimento de dados pendente)
processado_erro.loc[
    processado_erro['CAMPO'].str.contains('Enriquecimento de dados pendente', case=False),
    'RESPONSÁVEL'] = 'Central de Emissão'

# Realozação - Zupper
processado_erro.loc[
    processado_erro['OBTS'].str.contains('ZUPPER'),
    'RESPONSÁVEL'] = 'Operações - ZUPPER'

# Realocação - KONTRIP
processado_erro.loc[
    processado_erro['OBTS'].str.contains('KONTRIP'),
    'RESPONSÁVEL'] = 'Operações - KONTRIP'

# Realocação - INOVENTS
processado_erro.loc[
    processado_erro['Canal de Vendas'].str.contains('INOVENTS'),
    'RESPONSÁVEL'] = 'INOVENTS'

#---Preenchimento de Empresa---

# Realocações - EMPRESA - KONTIK BUSINESS TRAVEL
processado_erro.loc[
    processado_erro['RESPONSÁVEL'].str.contains('Operações - CORP') | processado_erro['RESPONSÁVEL'].str.contains('Operações - FCM') |
    processado_erro['RESPONSÁVEL'].str.contains('Suporte KCS'),
    'EMPRESA'] = 'KONTIK BUSINESS TRAVEL'

# Realocações - EMPRESA - ZUPPER VIAGENS
processado_erro.loc[
    processado_erro['OBTS'].str.contains('ZUPPER'),
    'EMPRESA'] = 'ZUPPER VIAGENS'

# Realocações - EMPRESA - INOVENTS
processado_erro.loc[
    processado_erro['RESPONSÁVEL'].str.contains('INOVENTS'),
    'EMPRESA'] = 'INOVENTS'

# Realocações - EMPRESA - GRUPO KONTIK
processado_erro.loc[
    processado_erro['RESPONSÁVEL'].str.contains('Central de Emissão') | processado_erro['RESPONSÁVEL'].str.contains('Conciliação aérea') |
    processado_erro['RESPONSÁVEL'].str.contains('Suporte Benner'),
    'EMPRESA'] = 'GRUPO KONTIK'

# Realocações - EMPRESA - KONTRIP VIAGENS
processado_erro.loc[
    processado_erro['OBTS'].str.contains('KONTRIP'),
    'EMPRESA'] = 'KONTRIP VIAGENS'

# Realocações - EMPRESA - K-CLUB
processado_erro.loc[
    processado_erro['RESPONSÁVEL'].str.contains('K-CLUB'),
    'EMPRESA'] = 'KTK'


#---Informativos---
# Verificação de duplicidade (Gover)
# Verifica se há mais de 10 ocorrências do mesmo localizador - GOVER
contagem_loc = processado_erro['Localizadora'].map(processado_erro['Localizadora'].value_counts())
mask_gover_dup = (contagem_loc > 10) & processado_erro['OBTS'].str.contains('GOVER', na=False)

processado_erro.loc[mask_gover_dup, ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'RESPONSÁVEL', 'CATEGORIA DE ERRO']] = [
    'Falha no processo de integração', 'Vendas duplicadas', 'Sistema', 'Suporte Benner', 'Sistêmico'
]

dup_gover = processado_erro[mask_gover_dup]
if not dup_gover.empty:
    primeiro = dup_gover.iloc[0]
    print(f'\033[1;31m-Verifique o localizador "{primeiro["Localizadora"]}"'
          f' e requisição "{primeiro["Requisição"]}",'
          f' feitas pelo consultor "{primeiro["Agente Emissão"]}"\033[m')
# Verifica se há Erros não identificados - Campo - Não Identificado
print(f'\033[1;33m- Identificamos {qtd} novos erros não categorizados.\033[m') if (qtd := len(processado_erro[processado_erro["CAMPO"] == "Não identificado"])) > 1 else None

# Salvar o arquivo original
processado_erro.to_excel(data_path + 'Processado Erro.xlsx', index=False)
 
# Seleção de colunas
colunas = {
    'A': 'Handle PNR', 'B': 'Handle ACC', 'C': 'Sequencia', 'D': 'Data Inclusão',
    'E': 'Data Alteração', 'F': 'Aging Alteração', 'G': 'Aging Inclusão', 'H': 'Localizadora', 'I': 'OBT', 'J': 'Pax',
    'K': 'Agente Emissão', 'L': 'Agente Criação', 'M': 'Data Emissão', 'N': 'Requisição', 'O': 'Local Retirada',
    'P': 'Status Requisicao', 'Q': 'Forma Pagamento', 'R': 'Forma Recebimento',
    'S': 'Serviço', 'T': 'Cancelado', 'U': 'Grupo Empresarial', 'V': 'Cliente',
    'W': 'Cliente Fee POS', 'X': 'Fornecedor', 'Y': 'Bilhete','Z': 'Canal de Vendas',
    'AA': 'Codigo Evento','AB': 'Tarifa', 'AC': 'Taxa', 'AD': 'Outras Taxas',
    'AE': 'Taxa DU', 'AF': 'Taxa BR', 'AG': 'Taxa Extra', 'AH': 'Fee',
    'AI': 'Observação', 'AJ': 'Mensagem Erro', 'AK': 'OBTS', 'AL': 'CAMPO',
    'AM': 'ORIGEM DO ERRO', 'AN': 'TIPO DE ERRO', 'AO': 'CATEGORIA DE ERRO',
    'AP': 'EMPRESA','AQ': 'RESPONSÁVEL'
}
 
# ---Escrita do relatório na planilha Excel---
# Grava o cabeçalho e os dados linha a linha na aba 'Processado Erro - BASE' do Dash.
# Inserção de cabeçalho
for col, nome in colunas.items():
    relatorio_base[col + '1'] = nome
 
# Inserção de informações
for row in range(len(processado_erro)):
    for col, nome in colunas.items():
        relatorio_base[col + str(row + 2)] = processado_erro[nome].iloc[row]
 
# ---Personalização visual da planilha---
# Aplica cores de cabeçalho (roxo para colunas base, verde para colunas calculadas),
# alinhamento à esquerda em todas as células e destaca em vermelho os agings críticos.
def personalizacao(relatorio):
    # Definindo cores para o cabeçalho
    colunas_padrao = PatternFill(start_color="591F6A", end_color="591F6A", fill_type="solid")  # ROXO
    colunas_adicionais = PatternFill(start_color="18F194", end_color="18F194", fill_type="solid")  # VERDE
    cor_da_fonte_1 = Font(color="18F194", bold=True) # cor forte VERDE
    cor_da_fonte_2 = Font(color="591F6A", bold=True) # cor fonte ROXO
    
    # Alinhamento à esquerda
    alinhamento_esquerda = Alignment(horizontal="left")

    # Aplicar a formatação aos cabeçalhos
    for coluna in range(1, 44):  # Colunas de A até AS
        letra_coluna = relatorio.cell(row=1, column=coluna).column_letter
        cabecalho = relatorio[letra_coluna + '1']
        cabecalho.fill = colunas_padrao
        cabecalho.font = cor_da_fonte_1
        # cabecalho.value = cabecalho.value.upper()  # Convertendo para maiúsculo
        
        if letra_coluna in ['F','G','AK','AL','AM', 'AN','AO','AP','AQ']:
            cabecalho.fill = colunas_adicionais
            cabecalho.font = cor_da_fonte_2
    
    # Alinhamento à esquerda nas células do corpo do relatório
    for row in relatorio.iter_rows(min_row=2, min_col=1, max_col=43):  # Ajuste para percorrer todas as colunas
        for cell in row:
            cell.alignment = alinhamento_esquerda  # Alinhando células à esquerda

    # Definindo cores para aging de dias
    criticos = ['16 a 23 dias', '24 a 31 dias', '31 dias ou +']
    for row in relatorio.iter_rows(min_row=2, min_col=6, max_col=7):
        if row[0].value in criticos:
            row[0].font = Font(color="FF0000")
        if row[1].value in criticos:
            row[1].font = Font(color="FF0000")

personalizacao(relatorio_base)

# Salvar relatorio base
dash.save(data_path + 'Relatorio - Dash.xlsx')

# ---Geração de relatórios individuais por empresa---
# Lê o relatório consolidado, lista as empresas únicas e cria um arquivo Excel
# separado para cada uma delas com a mesma formatação visual da planilha principal.
relatorio_dash = pd.read_excel(data_path + 'Relatorio - Dash.xlsx', sheet_name='Processado Erro - BASE')
# Obter lista de empresas
empresas = relatorio_dash['EMPRESA'].unique()

# Defina o caminho do diretório onde os arquivos serão salvos
output_folder = data_path + 'EMPRESAS'

# Verifique se o diretório existe; se não, crie-o
if os.path.exists(output_folder):
    shutil.rmtree(output_folder)  # Remove tudo: arquivos + subpastas

# Cria a pasta novamente, vazia
os.makedirs(output_folder, exist_ok=True)

for empresa in empresas:
    # Criar uma nova planilha para cada empresa
    df_empresa = relatorio_dash[relatorio_dash['EMPRESA'] == empresa]
    output_file = os.path.join(output_folder, f'Relatorio - {empresa}.xlsx')

    # Salvando o DataFrame da empresa em um arquivo Excel
    df_empresa.to_excel(output_file, index=False)

    # Abrindo o arquivo Excel para aplicar formatação
    dash = load_workbook(output_file)
    aba_relatorio = dash.active  # Pegando a primeira aba
    
    # Aplicando a personalização
    personalizacao(aba_relatorio)
    
    # Salvando as mudanças no arquivo Excel
    dash.save(output_file)

    print(f'Relatorio - {empresa} - Salvo com sucesso!')
print()
print('\033[1;32m-Relatório Processado Erro Gerado com sucesso!\033[m')
