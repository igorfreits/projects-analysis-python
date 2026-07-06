import pandas as pd
import os
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# --- Configuração ---
try:
    usuario = os.getlogin()
except OSError:
    usuario = os.environ.get('USERNAME') or os.environ.get('USER', 'usuario')

data_path = Path(f'C:/Users/{usuario}/Desktop/DOCS/data-analysis-python/Integration-Quality')
relatorio_path = data_path / 'Relatorio - Dash.xlsx'
base_path = data_path / 'Base.xlsx'

# Colunas que serão gravadas nas abas de saída (Base.xlsx)
COLUNAS_SAIDA = [
    'Status', 'Handle PNR', 'Handle ACC', 'Localizadora', 'Status Requisicao',
    'OBTS', 'Grupo Empresarial', 'Serviço', 'Mensagem Erro', 'TIPO DE ERRO',
    'EMPRESA', 'CATEGORIA DE ERRO', 'RESPONSÁVEL', 'Data Inclusão',
]

# --- Leitura ---
def ler_aba_excel(path: Path, sheet_name: str) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name=sheet_name)
    except FileNotFoundError:
        raise RuntimeError(f'Arquivo não encontrado: {path}') from None
    except ValueError as e:
        raise RuntimeError(f'Aba "{sheet_name}" não encontrada em {path.name}: {e}') from None

relatorio_dash = ler_aba_excel(relatorio_path, 'Processado Erro - BASE')

try:
    base_xls = pd.ExcelFile(base_path)
except FileNotFoundError:
    raise RuntimeError(f'Arquivo não encontrado: {base_path}') from None

def parse_aba(xls: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    if sheet_name not in xls.sheet_names:
        raise RuntimeError(f'Aba "{sheet_name}" não encontrada em {base_path.name}')
    return xls.parse(sheet_name=sheet_name)

base_em_andamento = parse_aba(base_xls, 'Benner - Processado Erro 0')
base_resolvidos    = parse_aba(base_xls, 'Resolvidos')
base_xls.close()

# --- Validação de colunas obrigatórias ---
colunas_necessarias = {'Handle ACC'}
for nome, df in [
    ('Relatorio - Dash', relatorio_dash),
    ('Em Andamento',     base_em_andamento),
    ('Resolvidos',       base_resolvidos),
]:
    faltando = colunas_necessarias - set(df.columns)
    if faltando:
        raise RuntimeError(f'Colunas ausentes em "{nome}": {faltando}')

# --- Processamento ---
data_atual = pd.Timestamp.today().normalize()

# ---Classificação de status de cada registro---
# Compara o relatório de hoje com as bases históricas e marca cada Handle ACC como:
# "Novo" (aparece pela primeira vez), "Em Andamento" (já existia) ou "Resolvido" (já concluído).
base_novo = relatorio_dash.reindex(columns=COLUNAS_SAIDA, fill_value=pd.NA)
base_novo['Status'] = 'Novo'
base_novo.loc[base_novo['Handle ACC'].isin(base_em_andamento['Handle ACC']), 'Status'] = 'Em Andamento'
base_novo.loc[base_novo['Handle ACC'].isin(base_resolvidos['Handle ACC']),   'Status'] = 'Resolvido'

# Novos registros → adiciona em base_em_andamento como "Em Andamento"
novos_registros = base_novo[base_novo['Status'] == 'Novo'].copy()
novos_registros['Status'] = 'Em Andamento'
print(f'\033[94m- Casos novos:\033[0m {len(novos_registros)}')

base_em_andamento = pd.concat([base_em_andamento, novos_registros], ignore_index=True)

# ---Atualização da base em andamento---
# Registros que não aparecem mais no relatório de hoje são marcados como "Resolvido".
# Os que são novos entram na base; os resolvidos são movidos para a aba de histórico.
# Atualiza status: registros ausentes no relatório → Resolvido
status_atual = base_novo[['Handle ACC', 'Status']]
base_em_andamento = base_em_andamento.merge(status_atual, on='Handle ACC', how='left', suffixes=('', '_novo'))
base_em_andamento['Status'] = base_em_andamento['Status_novo'].fillna('Resolvido')
base_em_andamento.drop(columns=['Status_novo'], inplace=True)

# Move registros resolvidos para base_resolvidos
registros_resolvidos = base_em_andamento[base_em_andamento['Status'] == 'Resolvido'].copy()
registros_resolvidos['Data de Conclusão'] = data_atual.date()

print(f'\033[94m- Total Processado Erro Hoje:\033[0m {len(relatorio_dash)}')
print(f'\033[94m- Casos resolvidos hoje:\033[0m {len(registros_resolvidos)}')

base_resolvidos = pd.concat([base_resolvidos, registros_resolvidos], ignore_index=True)

# Remove resolvidos da base em andamento
base_em_andamento = base_em_andamento[base_em_andamento['Status'] != 'Resolvido'].copy()

# Normaliza datas
for df, col in [(base_em_andamento, 'Data Inclusão'), (base_novo, 'Data Inclusão')]:
    df[col] = pd.to_datetime(df[col], errors='coerce').dt.date

base_resolvidos['Data de Conclusão'] = pd.to_datetime(
    base_resolvidos['Data de Conclusão'], errors='coerce'
)

# Mantém resolvidos do ano atual e do ano anterior
ano_limite = data_atual.year - 1
base_resolvidos = base_resolvidos[base_resolvidos['Data de Conclusão'].dt.year >= ano_limite].copy()

# ---Formatação visual do cabeçalho---
# Aplica fundo azul escuro, fonte branca em negrito, bordas finas e ajusta
# a largura das colunas automaticamente com base no conteúdo.
def formatar_cabecalho(ws):
    fill   = PatternFill('solid', fgColor='1F4E79')
    fonte  = Font(bold=True, color='FFFFFF', size=11)
    alinha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    borda_lado = Side(style='thin', color='BDD7EE')
    borda  = Border(left=borda_lado, right=borda_lado, top=borda_lado, bottom=borda_lado)

    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = fonte
        cell.alignment = alinha
        cell.border    = borda

    ws.row_dimensions[1].height = 30

    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

# ---Gravação das abas no Base.xlsx---
# Substitui (overlay) as três abas de controle com os dados atualizados
# e aplica a formatação de cabeçalho em cada uma delas.
try:
    with pd.ExcelWriter(base_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        base_novo.to_excel(writer,         sheet_name='Novo Arquivo',                 index=False)
        base_em_andamento.to_excel(writer, sheet_name='Benner - Processado Erro 0',   index=False)
        base_resolvidos.to_excel(writer,   sheet_name='Resolvidos',                   index=False)

        for nome_aba in ['Novo Arquivo', 'Benner - Processado Erro 0', 'Resolvidos']:
            formatar_cabecalho(writer.sheets[nome_aba])
except PermissionError:
    raise RuntimeError(f'Permissão negada ao salvar {base_path.name}. Feche o arquivo no Excel e tente novamente.') from None
except Exception as e:
    raise RuntimeError(f'Erro ao salvar {base_path.name}: {e}') from e

print(f'\033[1;92m\n- Atualização de status concluída com sucesso!\033[0m\n')
